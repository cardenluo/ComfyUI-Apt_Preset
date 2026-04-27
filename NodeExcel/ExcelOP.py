import os, re
from typing import Dict, List, Tuple

from PIL import Image as PILImage
from io import BytesIO
import csv





from ..main_unit import *


#------------------------------------------------------------
# 安全导入检查 -- 将导入语句修改为以下形式

try:
    import openpyxl
except ImportError:
    openpyxl = None
    print("Warning: openpyxl not installed, Excel-related nodes will not be available")

try:
    from openpyxl.drawing.image import Image as OpenpyxlImage
except ImportError:
    OpenpyxlImage = None
    print("Warning: openpyxl.drawing.image not available")

try:
    from openpyxl.utils import get_column_letter
except ImportError:
    get_column_letter = None
    print("Warning: openpyxl.utils.get_column_letter not available")


#------------------------------------------------------------


class excel_search_data:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "search_content": ("STRING", {"default": ""}),
                "search_mode": (["Precise_search", "Fuzzy_search"], {"default": "Precise_search"}),
            },
            "optional": {   } 
        }

    RETURN_TYPES = ("STRING", "INT", "INT")
    RETURN_NAMES = ("debug", "row", "col")
    FUNCTION = "search_data"
    CATEGORY = "Apt_Preset/prompt/excel"

    def IS_CHANGED(): return float("NaN")

    def search_data(self, excel_path, sheet_name, search_content, search_mode):
        try:
            if not os.path.exists(excel_path):
                return (f"Error: File does not exist at path: {excel_path}", None, None)
            if not os.access(excel_path, os.R_OK):
                return (f"Error: No read permission for file at path: {excel_path}", None, None)
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]

            results = []
            found_row = None
            found_col = None
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_value = cell.value if cell.value is not None else ""
                    cell_value_str = str(cell_value)
                    if (search_mode == "Precise_search" and cell_value_str == search_content) or \
                        (search_mode == "Fuzzy_search" and search_content in cell_value_str):
                        results.append(f"{sheet_name}|{row}|{col}|{cell_value}")
                        found_row = row
                        found_col = col

            workbook.close()
            del workbook
            if not results:
                return ("No results found.", None, None)
            return ("\n".join(results), found_row, found_col)
        except Exception as e:
            return (f"Error: {str(e)}", None, None)


class excel_row_diff:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "col_data": ("INT", {"default": 1, "min": 1, "step": 1}),
                "col_finish": ("INT", {"default": 1, "min": 1, "step": 1}),
            },
            "optional": {} 
        }

    RETURN_TYPES = ("INT", "INT", "INT")
    RETURN_NAMES = ("count", "count_data", "count_finish")
    FUNCTION = "excel_row_diff"
    CATEGORY = "Apt_Preset/prompt/excel"
    DESCRIPTION = """
    - col_data=2: 统计第2列,从上到下连续非空单元格总数count1
    - col_finish=3: 统计第3列,从上到下连续非空单元格总数count2
    - 输出未对齐的数量:count_data-count_finish
    """

    def IS_CHANGED(cls): 
        return float("NaN")

    def excel_row_diff(self, excel_path, sheet_name, col_data, col_finish):
        try:
            if not os.path.exists(excel_path):
                raise Exception(f"Error: File does not exist at path: {excel_path}")

            if not os.access(excel_path, os.R_OK):
                raise Exception(f"Error: No read permission for file at path: {excel_path}")

            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]

            def count_cells(col_index):
                if col_index == 0:  # 跳过无效列
                    return 0
                count = 0
                for row in range(1, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row, column=col_index).value
                    if cell_value is not None:
                        count += 1
                    else:
                        break
                return count

            count1 = count_cells(col_data)
            count2 = count_cells(col_finish)

            result = abs(count1 - count2)

            workbook.close()
            del workbook

            return (result, count1, count2)

        except Exception as e:
            raise Exception(f"Error: {str(e)}")


class excel_column_diff:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row_data": ("INT", {"default": 1, "min": 1, "step": 1}),
                "row_finish": ("INT", {"default": 1, "min": 1, "step": 1}),
            },
            "optional": {}
        }

    RETURN_TYPES = ("INT", "INT", "INT")
    RETURN_NAMES = ("count", "count_data", "count_finish")
    FUNCTION = "excel_column_diff"
    CATEGORY = "Apt_Preset/prompt/excel"
    DESCRIPTION = """
    - row_data=2: 统计第2行,从左到右连续非空单元格总数count1
    - row_finish=3: 统计第3行,从左到右连续非空单元格总数count2
    - 输出未对齐的数量:count_data-count_finish
    """

    def IS_CHANGED(cls):
        return float("NaN")

    def excel_column_diff(self, excel_path, sheet_name, row_data, row_finish):
        try:
            if not os.path.exists(excel_path):
                raise Exception(f"Error: File does not exist at path: {excel_path}")

            if not os.access(excel_path, os.R_OK):
                raise Exception(f"Error: No read permission for file at path: {excel_path}")

            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]

            def count_cells(row_index):
                if row_index == 0:  # 跳过无效行
                    return 0
                count = 0
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row_index, column=col).value
                    if cell_value is not None:
                        count += 1
                    else:
                        break
                return count

            count1 = count_cells(row_data)
            count2 = count_cells(row_finish)

            result = abs(count1 - count2)

            workbook.close()
            del workbook

            return (result, count1, count2)

        except Exception as e:
            raise Exception(f"Error: {str(e)}")


class excel_read:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row_start": ("INT", {"default": 0, "min": 0}),
                "row_end": ("INT", {"default": 3, "min": 1}),
                "col_start": ("INT", {"default": 0, "min": 0}),
                "col_end": ("INT", {"default": 4, "min": 1}),
            },
            "optional": {}
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("data",)
    FUNCTION = "excel_read"
    CATEGORY = "Apt_Preset/🚫Deprecated/🚫"
    DESCRIPTION = """
    - #excle最小单元行或列不能为0,将无效跳过
    - row_start=0, row_end=3: 单行输出: row=3    
    - row_start=1, row_end=3: 多行输出: row=1,2,3
    - row_start = row_end=1 : 行数相同,单行输出row=1
    - row_end=3, row_end=1 : 报错, 起始行必须小于=结束行
    """
    @classmethod
    def IS_CHANGED(cls):
        return float("NaN")

    def excel_read(self, excel_path, sheet_name, row_start, row_end, col_start, col_end):
        try:
            # 校验 start <= end 且 >= 0
            if row_start > row_end:
                raise Exception(f"Error: row_start ({row_start}) must be <= row_end ({row_end})!")
            if col_start > col_end:
                raise Exception(f"Error: col_start ({col_start}) must be <= col_end ({col_end})!")

            # 处理 row_start == 0 的情况：只取 row_end 行
            if row_start == 0:
                start_row = end_row = max(1, row_end)
            else:
                start_row, end_row = row_start, row_end

            # 处理 col_start == 0 的情况：只取 col_end 列
            if col_start == 0:
                start_col = end_col = max(1, col_end)
            else:
                start_col, end_col = col_start, col_end

            # 确保行列编号 ≥ 1
            start_row = max(1, start_row)
            end_row = max(1, end_row)
            start_col = max(1, start_col)
            end_col = max(1, end_col)

            # 打开 Excel 文件并读取数据
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]

            output_lines = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                output_lines.append("|".join(row_data))

            workbook.close()
            del workbook

            return ("\n".join(output_lines),)

        except Exception as e:
            return (f"Error: {str(e)}",)


class excel_write_data:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row_start": ("INT", {"default": 0, "min": 0}),
                "row_end": ("INT", {"default": 3, "min": 1}),
                "col_start": ("INT", {"default": 0, "min": 0}),
                "col_end": ("INT", {"default": 5, "min": 1}),
                "data": ("STRING", {"default": "", "multiline": True}),
            },
            "optional": {}
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("debug",)
    FUNCTION = "write_data"
    CATEGORY = "Apt_Preset/🚫Deprecated/🚫"
    DESCRIPTION = """
    - 示例: data 输入如下数据
    - 1 | 2 | 3 | 4
    - a | b | c | d
    - row_start=2, row_end=3  
    - col_start=2, col_end=5
    - 数据会依次填入第2行第2列 到 第3行第5列
    - 
    - row_start=0, row_end=3  
    - col_start=0, col_end=5
    - 数据会只填入到第3行第5列的单元格
    """

    @classmethod
    def IS_CHANGED(cls):
        return float("NaN")

    def write_data(self, excel_path, sheet_name, row_start, row_end, col_start, col_end, data):
        try:
            # 校验 start <= end 且 >= 0
            if row_start > row_end:
                return (f"Error: row_start ({row_start}) must be <= row_end ({row_end})",)
            if col_start > col_end:
                return (f"Error: col_start ({col_start}) must be <= col_end ({col_end})",)

            # 处理 start == 0 的情况：只写 end 行/列
            if row_start == 0:
                start_row = end_row = max(1, row_end)
            else:
                start_row, end_row = row_start, row_end

            if col_start == 0:
                start_col = end_col = max(1, col_end)
            else:
                start_col, end_col = col_start, col_end

            # 确保最小值为1（兼容 Excel）
            start_row = max(1, start_row)
            end_row = max(1, end_row)
            start_col = max(1, start_col)
            end_col = max(1, end_col)

            # 文件存在性及权限检查
            if not os.path.exists(excel_path):
                return (f"Error: File does not exist at path: {excel_path}",)
            if not os.access(excel_path, os.W_OK):
                return (f"Error: No write permission for file at path: {excel_path}",)

            # 加载工作簿和工作表
            workbook = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
            sheet = workbook[sheet_name]

            # 解析输入数据
            data_lines = data.strip().split("\n")

            # 写入数据
            for row_index, line in enumerate(data_lines, start=start_row):
                if row_index > end_row:
                    break
                cell_values = line.split("|")
                for col_index, cell_value in enumerate(cell_values, start=start_col):
                    if col_index > end_col:
                        break
                    if cell_value.strip():
                        sheet.cell(row=row_index, column=col_index).value = cell_value.strip()

            # 保存并关闭工作簿
            workbook.save(excel_path)
            workbook.close()
            del workbook

            return ("Data written successfully!",)

        except PermissionError as pe:
            return (f"Permission Error: {str(pe)}",)
        except Exception as e:
            return (f"Error: {str(e)}",)


class excel_insert_image:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row_start": ("INT", {"default": 0, "min": 0}),
                "row_end": ("INT", {"default": 1, "min": 1}),
                "col_start": ("INT", {"default": 0, "min": 0}),
                "col_end": ("INT", {"default": 1, "min": 1}),
                "img_height": ("INT", {"default": 256, "max": 2048, "min": 64}),
                "image_path": ("STRING", {"default": "image_file_path.png"}),
            },
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("debug",)
    FUNCTION = "write_image"
    CATEGORY = "Apt_Preset/🚫Deprecated/🚫"
    DESCRIPTION = """
    - 示例: 图片输入如下
    - row_start=2, row_end=3  
    - col_start=2, col_end=5
    - 图片会依次填入第2行第2列 到 第3行第5列
    - 
    - row_start=0, row_end=3  
    - col_start=0, col_end=5
    - 图片会只插入到第3行第5列的单元格
    """
    @classmethod
    def IS_CHANGED(cls):
        return float("NaN")

    def write_image(self, excel_path, sheet_name, row_start, row_end, col_start, col_end, image_path,img_height):
        try:
            # 校验 start <= end 且 >= 0
            if row_start > row_end:
                return (f"Error: row_start ({row_start}) must be <= row_end ({row_end})",)
            if col_start > col_end:
                return (f"Error: col_start ({col_start}) must be <= col_end ({col_end})",)

            # 处理 start == 0：只插入 end 所在行/列
            if row_start == 0:
                start_row = end_row = max(1, row_end)
            else:
                start_row, end_row = row_start, row_end

            if col_start == 0:
                start_col = end_col = max(1, col_end)
            else:
                start_col, end_col = col_start, col_end

            # 确保最小值为 1（兼容 Excel）
            start_row = max(1, start_row)
            end_row = max(1, end_row)
            start_col = max(1, start_col)
            end_col = max(1, end_col)

            # 文件存在性及权限检查
            if not os.path.exists(excel_path):
                return (f"Error: Excel file does not exist at path: {excel_path}",)
            if not os.access(excel_path, os.W_OK):
                return (f"Error: No write permission for Excel file at path: {excel_path}",)
            if not os.path.exists(image_path):
                return (f"Error: Image file does not exist at path: {image_path}",)
            if not os.access(image_path, os.R_OK):
                return (f"Error: No read permission for image file at path: {image_path}",)

            # 加载工作簿和工作表
            workbook = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
            sheet = workbook[sheet_name]

            # 插入图片的目标位置（仅使用 start 坐标）
            target_row = start_row
            target_col = start_col
            cell_address = get_column_letter(target_col) + str(target_row)

            # 打开图片并按比例缩放（高度固定为256像素，宽度按比例计算）
            with PILImage.open(image_path) as img:
                width, height = img.size
                scale = img_height / height
                target_width = int(width * scale)
                resized_img = img.resize((target_width, img_height), PILImage.LANCZOS)

                # 转换为字节流供 openpyxl 使用
                img_byte_arr = BytesIO()
                resized_img.save(img_byte_arr, format=img.format)
                openpyxl_img = OpenpyxlImage(img_byte_arr)

            # 调整单元格尺寸以适应图片
            column_letter = get_column_letter(target_col)
            
            # 设置列宽（按像素转磅的近似公式：1像素 ≈ 0.75磅）
            sheet.column_dimensions[column_letter].width = target_width * 0.75 / 7
            
            # 设置行高为192磅（对应256像素）
            sheet.row_dimensions[target_row].height = 192

            # 插入图片（图片会自动以单元格左上角对齐）
            sheet.add_image(openpyxl_img, cell_address)

            # 保存并关闭工作簿
            workbook.save(excel_path)
            workbook.close()

            return ("Image inserted and scaled successfully!",)

        except PermissionError as pe:
            return (f"Permission Error: {str(pe)}",)
        except Exception as e:
            return (f"Error: {str(e)}",)
        







#------------------------------------------------------------建设中------------------------





class excel_qwen_font:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    EXCEL_DIR = os.path.join(BASE_DIR, "qwen_Image")

    @staticmethod
    def load_excel_data(excel_path: str) -> dict:
        """加载Excel文件，返回{第一列: 第二列}格式的字典"""
        if not os.path.exists(excel_path):
            return {}
            
        try:
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook.active
            data = {}
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2 and row[0] is not None:
                    key = str(row[0]).strip()
                    value = str(row[1]) if row[1] is not None else ""
                    data[key] = value
            
            workbook.close()
            return data
        except:
            return {}

    def single_replace(self, text, target, replacement):
        if not target or not replacement:
            return text
        target_clean = target.strip('"').strip()
        replacement_clean = replacement.strip('"').strip()
        return re.sub(re.escape(target_clean), replacement_clean, text)

    @classmethod
    def INPUT_TYPES(cls):
        # 加载数据
        effect_path = os.path.join(cls.EXCEL_DIR, "文字效果.xlsx")
        font_path = os.path.join(cls.EXCEL_DIR, "文字类型.xlsx")
        array_path = os.path.join(cls.EXCEL_DIR, "文字排版.xlsx")
        
        cls.effect_data = cls.load_excel_data(effect_path)
        cls.font_data = cls.load_excel_data(font_path)
        cls.array_data = cls.load_excel_data(array_path)
        
        # 确保字典不为空
        if not cls.effect_data:
            cls.effect_data = {"默认效果": "标准效果"}
        if not cls.font_data:
            cls.font_data = {"默认字体": "常规字体"}
        if not cls.array_data:
            cls.array_data = {"默认排版": "标准排版"}
        
        return {
            "required": {
                "prefix_text": ("STRING", {"default": "", "multiline": True, "label": "前缀文本"}),
                "text_effect": (list(cls.effect_data.keys()), {"label": "文字效果"}),
                "text_font": (list(cls.font_data.keys()), {"label": "文字类型"}),
                "text_array": (list(cls.array_data.keys()), {"label": "文字排版"}),
                "text_medium": ("STRING", {"default": "", "multiline": False, "label": "媒介文本"}),
                "text_show": ("STRING", {"default": "", "multiline": False, "label": "显示文本"}),
                "suffix_text": ("STRING", {"default": "", "multiline": False, "label": "后缀文本"}),
                "neg": ("STRING", {"default": "", "multiline": False, "label": "负面提示"}),
            },
        }

    RETURN_TYPES = ("STRING", "STRING")
    RETURN_NAMES = ("pos", "neg")
    FUNCTION = "execute"
    CATEGORY = "Apt_Preset/🚫Deprecated/🚫"

    def execute(self, text_effect, text_font, text_array, prefix_text, 
                text_medium, text_show, suffix_text, neg):
        # 获取对应的值
        effect_val = self.effect_data.get(text_effect, "")
        font_val = self.font_data.get(text_font, "")
        array_val = self.array_data.get(text_array, "")
        
        # 处理默认值
        current_medium = text_medium if text_medium else "图像上"
        current_font = font_val if font_val else "Arial 字体，经典西文字体，字形规整，笔画简洁"
        
        # 构建核心内容基础部分
        core_base = f"在{current_medium}，以{current_font}呈现出文字 \"{text_show}\""
        
        # 构建附加部分（排版和效果）
        additional_parts = []
        if array_val:
            additional_parts.append(f"排版方式为{array_val}")
        if effect_val:
            additional_parts.append(f"效果内容为{effect_val}")
        
        # 组合核心内容
        if additional_parts:
            core_content = f"{core_base}，{','.join(additional_parts)}"
        else:
            core_content = core_base
        
        # 构建完整内容
        parts = []
        if prefix_text:
            parts.append(prefix_text)
        
        parts.append(core_content)
        
        if suffix_text:
            parts.append(suffix_text)
        
        # 组合所有部分
        pos = "，".join(parts) if parts else "未选择任何参数"
        
        # 替换占位符
        replacements = [
            ("{text_show}", text_show),
            ("{text_medium}", text_medium),
            ("{prefix}", prefix_text),
            ("{suffix}", suffix_text),
            ("{text_array}", text_array),
            ("{text_effect}", text_effect)
        ]
        
        for placeholder, value in replacements:
            if value:
                pos = self.single_replace(pos, placeholder, value)

        return (pos, neg)
    


class excel_qwen_artistic:   
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    EXCEL_DIR = os.path.join(BASE_DIR, "qwen_Image")
    DEFAULT_FONT_DESC = "Arial 字体，经典西文字体，字形规整，笔画简洁"

    @staticmethod
    def load_excel_data(excel_path: str) -> dict:
        if not os.path.exists(excel_path):
            return {}
        try:
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook.active
            data = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2 and row[0] is not None:
                    key = str(row[0]).strip()
                    value = str(row[1]) if row[1] is not None else ""
                    data[key] = value
            workbook.close()
            return data
        except:
            return {}

    def single_replace(self, text, target, replacement):
        if not target or not replacement:
            return text
        target_clean = target.strip('"').strip()
        replacement_clean = replacement.strip('"').strip()
        return re.sub(re.escape(target_clean), replacement_clean, text)

    @classmethod
    def INPUT_TYPES(cls):
        preset_path = os.path.join(cls.EXCEL_DIR, "模版预设.xlsx")
        medium_path = os.path.join(cls.EXCEL_DIR, "文字载体.xlsx")
        toon_path = os.path.join(cls.EXCEL_DIR, "整体氛围.xlsx")
        font_path = os.path.join(cls.EXCEL_DIR, "文字类型.xlsx")
        effect_path = os.path.join(cls.EXCEL_DIR, "文字效果.xlsx")
        array_path = os.path.join(cls.EXCEL_DIR, "文字排版.xlsx")
        
        cls.sum_preset_data = cls.load_excel_data(preset_path)
        cls.text_medium_data = cls.load_excel_data(medium_path)
        cls.sum_toon_data = cls.load_excel_data(toon_path)
        cls.text_font_data = cls.load_excel_data(font_path)
        cls.text_effect_data = cls.load_excel_data(effect_path)
        cls.text_array_data = cls.load_excel_data(array_path)
        
        if not cls.sum_preset_data:
            cls.sum_preset_data = {"默认预设": ""}
        if not cls.text_medium_data:
            cls.text_medium_data = {"默认载体": "图像上"}
        if not cls.sum_toon_data:
            cls.sum_toon_data = {"默认氛围": "常规氛围"}
        if not cls.text_font_data:
            cls.text_font_data = {"默认字体": cls.DEFAULT_FONT_DESC}
        if not cls.text_effect_data:
            cls.text_effect_data = {"默认效果": "标准效果"}
        if not cls.text_array_data:
            cls.text_array_data = {"默认排版": "常规排版"}
        
        return {
            "required": {
                "sum_preset": (list(cls.sum_preset_data.keys()), {"label": "模版预设"}),
                "text_medium": (list(cls.text_medium_data.keys()), {"label": "背景类型"}),
                "text_array": (list(cls.text_array_data.keys()), {"label": "字体样式"}),
                "text_font": (list(cls.text_font_data.keys()), {"label": "字体类型"}),
                "text_effect": (list(cls.text_effect_data.keys()), {"label": "字体效果"}),
                "sum_toon": (list(cls.sum_toon_data.keys()), {"label": "整体氛围"}),
                "text_cn": ("STRING", {"default": "", "multiline": False, "label": "中文文本"}),
                "text_en": ("STRING", {"default": "", "multiline": False, "label": "英文文本"}),
                "object": ("STRING", {"default": "", "multiline": False, "label": "元素文本"}),
                "target": ("STRING", {"default": "", "multiline": False, "label": "目标文本"}),
            },
            "optional": {
                "custom": ("STRING", {"default": "", "multiline": True, "placeholder": "输入=编辑新模板"}),
            }
        }
        
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("pos",)
    FUNCTION = "execute"
    CATEGORY = "Apt_Preset/prompt"

    DESCRIPTION = """
    - 编辑新模板，重新排版，可替换代码：
    "{text_medium}": medium_val,
    "{text_array}": array_val,
    "{text_font}": current_font,
    "{text_effect}": effect_val,
    "{sum_toon}": toon_val,
    "{text_cn}": text_cn,
    "{text_en}": text_en,
    "{object}": object,
    "{target}": target,
    "{text}": text_show

    - 举例：
    输入："文字的特点是 {text_font}，{sum_toon}" ，两个可替换内容，选择了 "宋体"，"蓝色氛围"
    输出结果："文字的特点是 宋体，蓝色氛围"

    """


    def execute(self, sum_preset, text_medium, text_array, text_font, text_effect, sum_toon, 
                text_cn="", text_en="", object="", target="", custom=""):
        text_show = text_cn.strip() if text_cn.strip() else text_en.strip() if text_en.strip() else "{text}"
        
        # 获取medium_val，确保有默认值"图像上"
        medium_val = self.text_medium_data.get(text_medium, "在图像上")
        if not medium_val.strip():  # 检查是否为空
            medium_val = "在图像上"
            
        font_val = self.text_font_data.get(text_font, "")
        array_val = self.text_array_data.get(text_array, "")
        effect_val = self.text_effect_data.get(text_effect, "")
        toon_val = self.sum_toon_data.get(sum_toon, "")  # 获取sum_toon对应的第二列内容
        current_font = font_val if font_val else self.DEFAULT_FONT_DESC
        
        # 检查是否有自定义模板输入，如果有且非空则使用自定义模板，否则使用预设模板
        if custom and custom.strip():
            preset_val = custom.strip()
            sum_preset_empty = False  # 自定义模板视为非空预设
        else:
            preset_val = self.sum_preset_data.get(sum_preset, "").strip()
            sum_preset_empty = (not preset_val)
        
        if sum_preset_empty:
            # 构建核心部分
            core_base = f"{medium_val}，以{current_font}呈现出文字\"{text_show}\""
            additional_parts = []
            
            # 添加排版和效果
            if array_val:
                additional_parts.append(f"排版方式为{array_val}")
            if effect_val:
                additional_parts.append(f"字体效果为{effect_val}")
            
            # 确保sum_toon内容被添加到最后
            if toon_val:
                additional_parts.append(f"整体的{toon_val}")
            
            # 组合结果
            if additional_parts:
                result = f"{core_base}，{','.join(additional_parts)}"
            else:
                result = core_base
                
            return (result,)
        
        # 处理有预设的情况（包括自定义模板）
        replacements = {
            "{text_medium}": medium_val,
            "{text_array}": array_val,
            "{text_font}": current_font,
            "{text_effect}": effect_val,
            "{sum_toon}": toon_val,
            "{text_cn}": text_cn,
            "{text_en}": text_en,
            "{object}": object,
            "{target}": target,
            "{text}": text_show
        }
        
        preset_content = preset_val
        for placeholder, value in replacements.items():
            if value:
                preset_content = self.single_replace(preset_content, placeholder, value)
        
        # 已移除：如果预设中没有包含sum_toon，不再自动添加到最后
        
        return (preset_content,)




class excel_VedioPrompt:   
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    EXCEL_PATH = os.path.join(BASE_DIR, "video", "视频提示词.xlsx")

    @staticmethod
    def load_excel_sheet(excel_path: str, sheet_name: str) -> dict:
        if not os.path.exists(excel_path):
            return {}
        try:
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            if sheet_name not in workbook.sheetnames:
                workbook.close()
                return {}
                
            sheet = workbook[sheet_name]
            data = {}
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or all(cell is None for cell in row):
                    continue
                    
                if row[0] is None:
                    continue
                    
                key = str(row[0]).strip()
                if key.startswith('#'):
                    continue
                    
                chinese_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                english_val = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                
                # 只在键是默认模板时才设置默认值
                if key == "默认模板" and not chinese_val and not english_val:
                    chinese_val = ""
                    english_val = ""
                    
                data[key] = {
                    'chinese': chinese_val,
                    'english': english_val
                }
            
            workbook.close()
            return data
        except Exception as e:
            print(f"加载表单 {sheet_name} 时出错: {e}")
            return {}

    def single_replace(self, text, target, replacement):
        if not target or not replacement:
            return text
        target_clean = target.strip('"').strip()
        replacement_clean = replacement.strip('"').strip()
        return re.sub(re.escape(target_clean), replacement_clean, text)

    @classmethod
    def INPUT_TYPES(cls):
        # 加载所有数据
        cls.template_data = cls.load_excel_sheet(cls.EXCEL_PATH, "模板")
        cls.style_data = cls.load_excel_sheet(cls.EXCEL_PATH, "风格")
        cls.scene_data = cls.load_excel_sheet(cls.EXCEL_PATH, "场景")
        cls.light_data = cls.load_excel_sheet(cls.EXCEL_PATH, "环境光")
        cls.camera_data = cls.load_excel_sheet(cls.EXCEL_PATH, "镜头")
        cls.atmosphere_data = cls.load_excel_sheet(cls.EXCEL_PATH, "氛围")
        cls.solar_term_data = cls.load_excel_sheet(cls.EXCEL_PATH, "节气")
        
        # 为各数据设置默认值
        if not cls.style_data:
            cls.style_data = {"默认风格": {'chinese': "写实风格", 'english': "realistic style"}}
        if not cls.light_data:
            cls.light_data = {"默认光线": {'chinese': "自然光", 'english': "natural light"}}
        if not cls.scene_data:
            cls.scene_data = {"默认场景": {'chinese': "室内场景", 'english': "indoor scene"}}
        if not cls.camera_data:
            cls.camera_data = {"默认镜头": {'chinese': "标准镜头", 'english': "standard lens"}}
        if not cls.atmosphere_data:
            cls.atmosphere_data = {"默认氛围": {'chinese': "舒适氛围", 'english': "comfortable atmosphere"}}
        if not cls.solar_term_data:
            cls.solar_term_data = {"默认节气": {'chinese': "春季", 'english': "spring"}}
        
        # 仅在模板数据为空时才添加默认模板选项
        if not cls.template_data:
            cls.template_data = {"默认模板": {'chinese': "", 'english': ""}}
        
        return {
            "required": {
                "template": (list(cls.template_data.keys()), {"label": "模板"}),
                "roles":  ("STRING", {"default": "", "multiline": False, }),
                "style": (list(cls.style_data.keys()), {"label": "风格"}),
                "light": (list(cls.light_data.keys()), {"label": "环境光"}),
                "scene": (list(cls.scene_data.keys()), {"label": "场景"}),
                "camera": (list(cls.camera_data.keys()), {"label": "镜头"}),
                "atmosphere": (list(cls.atmosphere_data.keys()), {"label": "氛围"}),
                "solar_term": (list(cls.solar_term_data.keys()), {"label": "节气"}),
                "language": ("BOOLEAN", {"default": True, "label_on": "中文", "label_off": "英文"}),
                "object": ("STRING", {"default": "", "multiline": False, "label": "元素文本"}),
                "target": ("STRING", {"default": "", "multiline": False, "label": "目标文本"}),
 
            },
            "optional": {
                "custom": ("STRING", {"default": "", "multiline": True, "placeholder": "输入=编辑新模版"})
            }
        }
        
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("pos",)
    FUNCTION = "execute"
    CATEGORY = "Apt_Preset/prompt"

    DESCRIPTION = """
    - 编辑新模板，重新排版，可替换代码：
    "{roles}": roles_val,
    "{style}": style_val,
    "{light}": light_val,
    "{scene}": scene_val,
    "{camera}": camera_val,
    "{atmosphere}": atmosphere_val,
    "{solar_term}": solar_term_val,
    "{object}": object,
    "{target}": target

    - 举例：
    输入："{light}，女孩在{scene}玩耍" ，两个可替换内容，选择了 "白天"，"草坪"
    输出结果："白天，女孩在草坪玩耍"

    """


    def execute(self, style, light, scene, camera, atmosphere, solar_term, template,
                language=True, object="", target="", custom="", roles=""):
        if not roles:
            roles = ""
            
        lang = 'chinese' if language else 'english'
     
        style_val = self.style_data.get(style, {}).get(lang, "默认风格" if language else "default style")
        light_val = self.light_data.get(light, {}).get(lang, "自然光" if language else "natural light")
        scene_val = self.scene_data.get(scene, {}).get(lang, "室内场景" if language else "indoor scene")
        camera_val = self.camera_data.get(camera, {}).get(lang, "标准镜头" if language else "standard lens")
        atmosphere_val = self.atmosphere_data.get(atmosphere, {}).get(lang, "舒适氛围" if language else "comfortable atmosphere")
        solar_term_val = self.solar_term_data.get(solar_term, {}).get(lang, "春季" if language else "spring")
        roles_val = roles.strip()  # 获取角色输入值
        custom_val = custom.strip()  # 获取自定义内容
        
        # 检查是否使用自定义模板
        if custom_val and not (self.template_data.get(template, {}).get(lang, "").strip()):
            # 当模板为空且custom有内容时，先构建默认公式再追加custom内容
            parts = []
            if roles_val:
                parts.append(roles_val)
            parts.extend([style_val, light_val, scene_val, camera_val, atmosphere_val, solar_term_val])
            
            # 过滤空值
            parts = [part for part in parts if part]
            
            separator = "，" if language else ", "
            result = separator.join(parts)
            
            # 追加custom内容
            if result:  # 如果已有内容，先加分隔符
                result += separator + custom_val
            else:  # 如果没有其他内容，直接使用custom内容
                result = custom_val
                
            return (result,)
        elif custom_val:
            # 当模板不为空但有custom内容时，直接使用custom内容
            template_val = custom_val
            use_formula = False
        else:
            # 获取选中的模板值
            template_val = self.template_data.get(template, {}).get(lang, "").strip()
            # 如果模板值为空，则使用公式
            use_formula = (not template_val)
        
        # 如果需要使用公式
        if use_formula:
            # 构建部分列表，包含角色（如果有输入）
            parts = []
            if roles_val:
                parts.append(roles_val)
            parts.extend([style_val, light_val, scene_val, camera_val, atmosphere_val, solar_term_val])
            
            # 过滤空值
            parts = [part for part in parts if part]
            
            separator = "，" if language else ", "
            result = separator.join(parts)
            return (result,)
        
        # 否则使用模板替换
        replacements = {
            "{roles}": roles_val,
            "{style}": style_val,
            "{light}": light_val,
            "{scene}": scene_val,
            "{camera}": camera_val,
            "{atmosphere}": atmosphere_val,
            "{solar_term}": solar_term_val,
            "{object}": object,
            "{target}": target
        }
        
        template_content = template_val
        for placeholder, value in replacements.items():
            if value:
                template_content = self.single_replace(template_content, placeholder, value)
        
        return (template_content,)




class excel_roles:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    EXCEL_PATH = os.path.join(BASE_DIR, "video", "角色.xlsx")

    @staticmethod
    def load_excel_sheet(excel_path: str, sheet_name: str) -> dict:
        if not os.path.exists(excel_path):
            return {}
        try:
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            if sheet_name not in workbook.sheetnames:
                workbook.close()
                return {}
                
            sheet = workbook[sheet_name]
            data = {}
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or all(cell is None for cell in row):
                    continue
                    
                if row[0] is None:
                    continue
                    
                key = str(row[0]).strip()
                if key.startswith('#'):
                    continue
                    
                chinese_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                english_val = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                
                # 只在键是默认模板时才设置默认值
                if key == "默认模板" and not chinese_val and not english_val:
                    # 默认模板，将所有的文本用逗号隔开，串联起来
                    chinese_val = "模版,角色,服饰,发型,皮肤,脸型,耳朵,眼睛,眉毛,鼻子,嘴巴,身材,胸部,动作,表情"
                    english_val = "template,role,clothing,hairstyle,skin,face shape,ears,eyes,eyebrows,nose,mouth,figure,chest,action,expression"
                    
                data[key] = {
                    'chinese': chinese_val,
                    'english': english_val
                }
            
            workbook.close()
            return data
        except Exception as e:
            print(f"加载表单 {sheet_name} 时出错: {e}")
            return {}

    def single_replace(self, text, target, replacement):
        if not target or not replacement:
            return text
        target_clean = target.strip('"').strip()
        replacement_clean = replacement.strip('"').strip()
        return re.sub(re.escape(target_clean), replacement_clean, text)

    @classmethod
    def INPUT_TYPES(cls):
        # 加载所有数据
        cls.template_data = cls.load_excel_sheet(cls.EXCEL_PATH, "模版")
        cls.role_data = cls.load_excel_sheet(cls.EXCEL_PATH, "角色")
        cls.clothing_data = cls.load_excel_sheet(cls.EXCEL_PATH, "服饰")
        cls.hairstyle_data = cls.load_excel_sheet(cls.EXCEL_PATH, "发型")
        cls.skin_data = cls.load_excel_sheet(cls.EXCEL_PATH, "皮肤")
        cls.face_shape_data = cls.load_excel_sheet(cls.EXCEL_PATH, "脸型")
        cls.ears_data = cls.load_excel_sheet(cls.EXCEL_PATH, "耳朵")
        cls.eyes_data = cls.load_excel_sheet(cls.EXCEL_PATH, "眼睛")
        cls.eyebrows_data = cls.load_excel_sheet(cls.EXCEL_PATH, "眉毛")
        cls.nose_data = cls.load_excel_sheet(cls.EXCEL_PATH, "鼻子")
        cls.mouth_data = cls.load_excel_sheet(cls.EXCEL_PATH, "嘴巴")
        cls.figure_data = cls.load_excel_sheet(cls.EXCEL_PATH, "身材")
        cls.chest_data = cls.load_excel_sheet(cls.EXCEL_PATH, "胸部")
        cls.action_data = cls.load_excel_sheet(cls.EXCEL_PATH, "动作")
        cls.expression_data = cls.load_excel_sheet(cls.EXCEL_PATH, "表情")
        
        # 为各数据设置默认值
        if not cls.role_data:
            cls.role_data = {"默认角色": {'chinese': "普通人", 'english': "ordinary person"}}
        if not cls.clothing_data:
            cls.clothing_data = {"默认服饰": {'chinese': "日常服装", 'english': "everyday clothing"}}
        if not cls.hairstyle_data:
            cls.hairstyle_data = {"默认发型": {'chinese': "短发", 'english': "short hair"}}
        if not cls.skin_data:
            cls.skin_data = {"默认皮肤": {'chinese': "自然肤色", 'english': "natural skin tone"}}
        if not cls.face_shape_data:
            cls.face_shape_data = {"默认脸型": {'chinese': "圆形脸", 'english': "round face"}}
        if not cls.ears_data:
            cls.ears_data = {"默认耳朵": {'chinese': "普通耳朵", 'english': "normal ears"}}
        if not cls.eyes_data:
            cls.eyes_data = {"默认眼睛": {'chinese': "黑色眼睛", 'english': "black eyes"}}
        if not cls.eyebrows_data:
            cls.eyebrows_data = {"默认眉毛": {'chinese': "自然眉形", 'english': "natural eyebrows"}}
        if not cls.nose_data:
            cls.nose_data = {"默认鼻子": {'chinese': "中等鼻子", 'english': "average nose"}}
        if not cls.mouth_data:
            cls.mouth_data = {"默认嘴巴": {'chinese': "自然嘴唇", 'english': "natural lips"}}
        if not cls.figure_data:
            cls.figure_data = {"默认身材": {'chinese': "匀称身材", 'english': "well-proportioned figure"}}
        if not cls.chest_data:
            cls.chest_data = {"默认胸部": {'chinese': "适中胸部", 'english': "moderate chest"}}
        if not cls.action_data:
            cls.action_data = {"默认动作": {'chinese': "站立", 'english': "standing"}}
        if not cls.expression_data:
            cls.expression_data = {"默认表情": {'chinese': "自然表情", 'english': "natural expression"}}
        
        # 仅在模板数据为空时才添加默认模板选项
        if not cls.template_data:
            cls.template_data = {"默认模板": {'chinese': "", 'english': ""}}
        
        return {
            "required": {
                "template": (list(cls.template_data.keys()), {"label": "模版"}),
                "role": (list(cls.role_data.keys()), {"label": "角色"}),
                "clothing": (list(cls.clothing_data.keys()), {"label": "服饰"}),
                "hairstyle": (list(cls.hairstyle_data.keys()), {"label": "发型"}),
                "skin": (list(cls.skin_data.keys()), {"label": "皮肤"}),
                "face_shape": (list(cls.face_shape_data.keys()), {"label": "脸型"}),
                "ears": (list(cls.ears_data.keys()), {"label": "耳朵"}),
                "eyes": (list(cls.eyes_data.keys()), {"label": "眼睛"}),
                "eyebrows": (list(cls.eyebrows_data.keys()), {"label": "眉毛"}),
                "nose": (list(cls.nose_data.keys()), {"label": "鼻子"}),
                "mouth": (list(cls.mouth_data.keys()), {"label": "嘴巴"}),
                "figure": (list(cls.figure_data.keys()), {"label": "身材"}),
                "chest": (list(cls.chest_data.keys()), {"label": "胸部"}),
                "action": (list(cls.action_data.keys()), {"label": "动作"}),
                "expression": (list(cls.expression_data.keys()), {"label": "表情"}),
                "language": ("BOOLEAN", {"default": True, "label_on": "中文", "label_off": "英文"})
            },
            "optional": {
                "custom": ("STRING", {"default": "", "multiline": True, "placeholder": "输入=编辑新模版"})
            }
        }
        
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("pos",)
    FUNCTION = "execute"
    CATEGORY = "Apt_Preset/prompt"
    DESCRIPTION = """
    - 编辑新模板，重新排版，可替换代码：
    "{role}": "角色",
    "{clothing}": "服饰",
    "{hairstyle}": "发型",
    "{skin}": "肤色",
    "{faceshape}": "脸型",
    "{ears}": "耳朵",
    "{eyes}": "眼睛",
    "{eyebrows}": "眉毛",
    "{nose}": "鼻子",
    "{mouth}": "嘴巴",
    "{figure}": "身材",
    "{chest}": "胸部",
    "{action}": "动作",
    "{expression}": "表情"
    例如：
    "{role}显得很{expression}"   输出：警察显得很惊讶
    """


    def execute(self, template, role, clothing, hairstyle, skin, face_shape, ears, eyes,
                eyebrows, nose, mouth, figure, chest, action, expression, language=True,
                custom=""):
        lang = 'chinese' if language else 'english'
     
        template_val = self.template_data.get(template, {}).get(lang, "")
        role_val = self.role_data.get(role, {}).get(lang, "普通人" if language else "ordinary person")
        clothing_val = self.clothing_data.get(clothing, {}).get(lang, "日常服装" if language else "everyday clothing")
        hairstyle_val = self.hairstyle_data.get(hairstyle, {}).get(lang, "短发" if language else "short hair")
        skin_val = self.skin_data.get(skin, {}).get(lang, "自然肤色" if language else "natural skin tone")
        face_shape_val = self.face_shape_data.get(face_shape, {}).get(lang, "圆形脸" if language else "round face")
        ears_val = self.ears_data.get(ears, {}).get(lang, "普通耳朵" if language else "normal ears")
        eyes_val = self.eyes_data.get(eyes, {}).get(lang, "黑色眼睛" if language else "black eyes")
        eyebrows_val = self.eyebrows_data.get(eyebrows, {}).get(lang, "自然眉形" if language else "natural eyebrows")
        nose_val = self.nose_data.get(nose, {}).get(lang, "中等鼻子" if language else "average nose")
        mouth_val = self.mouth_data.get(mouth, {}).get(lang, "自然嘴唇" if language else "natural lips")
        figure_val = self.figure_data.get(figure, {}).get(lang, "匀称身材" if language else "well-proportioned figure")
        chest_val = self.chest_data.get(chest, {}).get(lang, "适中胸部" if language else "moderate chest")
        action_val = self.action_data.get(action, {}).get(lang, "站立" if language else "standing")
        expression_val = self.expression_data.get(expression, {}).get(lang, "自然表情" if language else "natural expression")
        
        custom_val = custom.strip()  # 获取自定义内容
        
        # 检查是否使用自定义模板
        if custom_val and not (self.template_data.get(template, {}).get(lang, "").strip()):
            # 当模板为空且custom有内容时，先构建默认公式再追加custom内容
            parts = [
                role_val, clothing_val, hairstyle_val, skin_val, face_shape_val,
                ears_val, eyes_val, eyebrows_val, nose_val, mouth_val,
                figure_val, chest_val, action_val, expression_val
            ]
            
            # 过滤空值
            parts = [part for part in parts if part]
            
            separator = "，" if language else ", "
            result = separator.join(parts)
            
            # 追加custom内容
            if result:  # 如果已有内容，先加分隔符
                result += separator + custom_val
            else:  # 如果没有其他内容，直接使用custom内容
                result = custom_val
                
            return (result,)
        elif custom_val:
            # 当模板不为空但有custom内容时，直接使用custom内容
            template_val = custom_val
            use_formula = False
        else:
            # 获取选中的模板值
            template_val = self.template_data.get(template, {}).get(lang, "").strip()
            # 如果模板值为空，则使用公式
            use_formula = (not template_val)
        
        # 如果需要使用公式
        if use_formula:
            # 构建部分列表
            parts = [
                role_val, clothing_val, hairstyle_val, skin_val, face_shape_val,
                ears_val, eyes_val, eyebrows_val, nose_val, mouth_val,
                figure_val, chest_val, action_val, expression_val
            ]
            
            # 过滤空值
            parts = [part for part in parts if part]
            
            separator = "，" if language else ", "
            result = separator.join(parts)
            return (result,)

        replacements = {
            "{role}": role_val,
            "{clothing}": clothing_val,
            "{hairstyle}": hairstyle_val,
            "{skin}": skin_val,
            "{faceshape}": face_shape_val,
            "{ears}": ears_val,
            "{eyes}": eyes_val,
            "{eyebrows}": eyebrows_val,
            "{nose}": nose_val,
            "{mouth}": mouth_val,
            "{figure}": figure_val,
            "{chest}": chest_val,
            "{action}": action_val,
            "{expression}": expression_val
        }
        
        template_content = template_val
        for placeholder, value in replacements.items():
            if value:
                template_content = self.single_replace(template_content, placeholder, value)
        
        if not template_content.strip():
            template_content = "None"
        return (template_content,)




class excel_insert_image_easy:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row": ("INT", {"default": 1, "min": 1}),
                "column": ("INT", {"default": 1, "min": 1}),
                "img_height": ("INT", {"default": 256, "max": 2048, "min": 64}),
                "image": ("IMAGE",),
            },
            "optional": {
                "debug_output": ("BOOLEAN", {"default": False, "label_on": "启用", "label_off": "禁用"}),
            },
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("debug",)
    FUNCTION = "write_image"
    CATEGORY = "Apt_Preset/prompt/excel"
    OUTPUT_NODE = True  # 标记为输出节点

    @classmethod
    def IS_CHANGED(cls):
        return float("NaN")

    def write_image(self, excel_path, sheet_name, row, column, img_height, image, debug_output=False):
        try:
            target_row = max(1, row)
            target_col = max(1, column)

            if not os.path.exists(excel_path):
                error_msg = f"Error: Excel file does not exist at path: {excel_path}"
                if debug_output:
                    print(error_msg)
                return (error_msg,)
            if not os.access(excel_path, os.W_OK):
                error_msg = f"Error: No write permission for Excel file at path: {excel_path}"
                if debug_output:
                    print(error_msg)
                return (error_msg,)

            workbook = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
            sheet = workbook[sheet_name]

            cell_address = get_column_letter(target_col) + str(target_row)

            if hasattr(image, 'cpu'):
                image_np = image.cpu().numpy()
            else:
                image_np = image

            if len(image_np.shape) == 4:
                image_np = image_np[0]

            if image_np.max() <= 1.0:
                image_np = (image_np * 255).astype('uint8')
            else:
                image_np = image_np.astype('uint8')

            pil_img = PILImage.fromarray(image_np)

            width, height = pil_img.size
            scale = img_height / height
            target_width = int(width * scale)
            resized_img = pil_img.resize((target_width, img_height), PILImage.LANCZOS)

            img_byte_arr = BytesIO()
            resized_img.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)
            openpyxl_img = OpenpyxlImage(img_byte_arr)

            column_letter = get_column_letter(target_col)
            
            sheet.column_dimensions[column_letter].width = target_width * 0.75 / 7
            
            sheet.row_dimensions[target_row].height = 192

            sheet.add_image(openpyxl_img, cell_address)

            workbook.save(excel_path)
            workbook.close()

            success_msg = "Image inserted and scaled successfully!"
            if debug_output:
                print(success_msg)
            return (success_msg,)

        except PermissionError as pe:
            error_msg = f"Permission Error: {str(pe)}"
            if debug_output:
                print(error_msg)
            return (error_msg,)
        except Exception as e:
            error_msg = f"Error: {str(e)}"
            if debug_output:
                print(error_msg)
            return (error_msg,)




class excel_read_easy:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row": ("INT", {"default": 1, "min": 1}),
                "column": ("INT", {"default": 1, "min": 1}),
            },
            "optional": {}
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("data",)
    FUNCTION = "excel_read"
    CATEGORY = "Apt_Preset/prompt/excel"

    @classmethod
    def IS_CHANGED(cls):
        return float("NaN")

    def excel_read(self, excel_path, sheet_name, row, column):
        try:
            target_row = max(1, row)
            target_col = max(1, column)

            if not os.path.exists(excel_path):
                return (f"Error: File does not exist at path: {excel_path}",)

            if not os.access(excel_path, os.R_OK):
                return (f"Error: No read permission for file at path: {excel_path}",)

            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]

            cell_value = sheet.cell(row=target_row, column=target_col).value
            result = str(cell_value) if cell_value is not None else ""

            workbook.close()
            del workbook

            return (result,)

        except Exception as e:
            return (f"Error: {str(e)}",)



class excel_write_data_easy:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row": ("INT", {"default": 1, "min": 1}),
                "column": ("INT", {"default": 1, "min": 1}),
                "data": ("STRING", {"forceInput": True, }),
            },
            "optional": {
            }
        }
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("debug",)
    FUNCTION = "write_data"
    CATEGORY = "Apt_Preset/prompt/excel"
    OUTPUT_NODE = True
    @classmethod
    def IS_CHANGED(cls):
        return float("NaN")
    def write_data(self, excel_path, sheet_name, row, column, data,):

        try:
            target_row = max(1, row)
            target_col = max(1, column)
            if not os.path.exists(excel_path):
                error_msg = f"Error: File does not exist at path: {excel_path}"

            if not os.access(excel_path, os.W_OK):
                error_msg = f"Error: No write permission for file at path: {excel_path}"

            workbook = openpyxl.load_workbook(excel_path)
            if sheet_name not in workbook.sheetnames:
                workbook.create_sheet(sheet_name)
            sheet = workbook[sheet_name]
            sheet.cell(row=target_row, column=target_col).value = None
            if data.strip():
                sheet.cell(row=target_row, column=target_col).value = data.strip()
                sheet.cell(row=target_row, column=target_col).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')

                success_msg = f"成功将数据整体写入行 {target_row} 列 {target_col}"
            else:
                success_msg = "未写入数据，输入内容为空"
            workbook.save(excel_path)
            workbook.close()

            return (success_msg,)
        except PermissionError as pe:
            error_msg = f"权限错误: {str(pe)}"

            return (error_msg,)
        except Exception as e:
            error_msg = f"错误: {str(e)}"
            return (error_msg,)
    



class excel_Prompter:  
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    EXCEL_DIR = os.path.join(BASE_DIR, "global")
    CONFIG_FILE = os.path.join(EXCEL_DIR, "config.xlsx")
    excel_data: Dict[str, Dict] = {}
    excel_files: List[str] = []
    
    @classmethod
    def load_config(cls) -> List[str]:
        if not os.path.exists(cls.CONFIG_FILE):
            print(f"配置文件不存在: {cls.CONFIG_FILE}")
            return []
        try:
            workbook = openpyxl.load_workbook(cls.CONFIG_FILE, read_only=True, data_only=True)
            sheet = workbook.active
            xlsx_files = []
            row = 1
            while True:
                cell_value = sheet.cell(row=row, column=2).value
                if cell_value is None:
                    break
                file_name = str(cell_value).strip()
                if file_name and file_name.lower().endswith('.xlsx'):
                    xlsx_files.append(file_name)
                row += 1
            workbook.close()
            cls.excel_files = [os.path.splitext(f)[0] for f in xlsx_files]
            return xlsx_files
        except Exception as e:
            print(f"读取配置文件失败: {e}")
            return []
    
    @classmethod
    def load_all_excels(cls) -> None:
        xlsx_files = cls.load_config()
        cls.excel_data = {}
        for file in xlsx_files:
            key = os.path.splitext(file)[0]
            file_path = os.path.join(cls.EXCEL_DIR, file)
            cls.excel_data[key] = cls.load_excel(file_path)
    
    @staticmethod
    def load_excel(excel_path: str) -> dict:
        if not os.path.exists(excel_path):
            print(f"Excel文件不存在: {excel_path}")
            return {"文件不存在": ["", "", ""]}
        try:
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook.active
            headers = next(sheet.iter_rows(values_only=True))
            id_col = headers.index('ID') if 'ID' in headers else 0
            pos_col = headers.index('Positive') if 'Positive' in headers else 1
            neg_col = headers.index('Negative') if 'Negative' in headers else 2
            help_col = headers.index('Help') if 'Help' in headers else 3
            data = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[id_col]:
                    pos_val = str(row[pos_col] or '') if len(row) > pos_col else ''
                    neg_val = str(row[neg_col] or '') if len(row) > neg_col else ''
                    help_val = str(row[help_col] or '') if len(row) > help_col else ''
                    data[row[id_col]] = [pos_val, neg_val, help_val]
            workbook.close()
            return data if data else {"无有效数据": ["", "", ""]}
        except Exception as e:
            print(f"加载Excel失败: {excel_path} - {e}")
            return {"加载失败": ["", "", ""]}
    
    @staticmethod
    def split_with_quotes(s):
        pattern = r'"([^"]*)"|\s*([^,]+)'
        matches = re.finditer(pattern, s)
        return [match.group(1) or match.group(2).strip() for match in matches if match.group(1) or match.group(2).strip()]
    
    def single_replace(self, text, target, replacement):
        if not target or not replacement:
            return text
        target_clean = target.strip('"').strip()
        replacement_clean = replacement.strip('"').strip()
        pattern = re.escape(target_clean)
        return re.sub(pattern, replacement_clean, text)
    
    def multi_replace(self, text, multi_targets, multi_replacements):
        if not multi_targets or not multi_replacements:
            return text
        targets = multi_targets.split('@')
        replacements = multi_replacements.split('@')
        min_len = min(len(targets), len(replacements))
        targets = targets[:min_len]
        replacements = replacements[:min_len]
        result = text
        for target, replacement in zip(targets, replacements):
            result = self.single_replace(result, target.strip(), replacement.strip())
        return result

    @classmethod
    def INPUT_TYPES(cls):
        cls.load_all_excels()
        input_config = {
            "required": {

            },
            "optional": {
                "language": ("BOOLEAN", {"default": True, "label_on": "英文", "label_off": "中文"}),
            }
        }
        for file_key in cls.excel_files:
            options = list(cls.excel_data[file_key].keys())
            options.insert(0, "None")
            input_config["required"][file_key] = (options,)
        return input_config
    
    RETURN_TYPES = ("STRING", )  
    RETURN_NAMES = ("pos", )  
    FUNCTION = "execute"
    CATEGORY = "Apt_Preset/prompt"

    def execute(self, language=True, **kwargs):
        excel_positives = []
        excel_negatives = []
        
        for file_key in self.excel_files:
            selected_value = kwargs.get(file_key)
            if selected_value == "None":
                continue
            if selected_value in self.excel_data[file_key]:
                excel_data = self.excel_data[file_key][selected_value]
                pos_parts = [part.strip() for part in self.split_with_quotes(excel_data[0]) if part.strip()]
                neg_parts = [part.strip() for part in self.split_with_quotes(excel_data[1]) if part.strip()]
                
                if pos_parts:
                    selected_pos = pos_parts[0]
                    excel_positives.append(selected_pos)
                if neg_parts:
                    selected_neg = neg_parts[0]
                    excel_negatives.append(selected_neg)
        
        positive_prompt = ", ".join(filter(None, excel_positives))
        negative_prompt = ", ".join(filter(None, excel_negatives))
        
        final_output = negative_prompt if language else positive_prompt
        
        if not final_output:
            final_output = "❌ No valid excel content"
        
        return (final_output, )




class excel_video_assistant:
 
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    CSV_PATH = os.path.join(BASE_DIR, "video", "video_assistant.csv")

    @staticmethod
    def load_csv_data(csv_path: str) -> dict:
        """加载 CSV 文件，返回分类到标题到内容的字典"""
        if not os.path.exists(csv_path):
            print(f"CSV 文件不存在：{csv_path}")
            return {}
        
        try:
            data = {}
            
            with open(csv_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                headers = next(reader, [])  # 读取表头
                
                # 确定列索引
                try:
                    idx_title = headers.index('标题')
                    idx_type = headers.index('类型')
                    idx_category = headers.index('分类')
                    idx_content = headers.index('内容')
                    idx_variables = headers.index('变量')
                except ValueError as e:
                    print(f"CSV 文件列名错误：{e}")
                    return {}
                
                for row in reader:
                    if len(row) > max(idx_title, idx_type, idx_category, idx_content):
                        item_type = str(row[idx_type]).strip() if idx_type < len(row) else ""
                        item_category = str(row[idx_category]).strip() if idx_category < len(row) else ""
                        title = str(row[idx_title]).strip() if idx_title < len(row) else ""
                        content = str(row[idx_content]) if idx_content < len(row) else ""
                        variables = str(row[idx_variables]) if idx_variables < len(row) else ""
                        
                        # 只处理"视频编辑"类型的数据
                        if item_type == "视频编辑" and title and item_category:
                            if item_category not in data:
                                data[item_category] = {}
                            data[item_category][title] = {
                                'content': content,
                                'variables': variables
                            }
            
            return data
        except Exception as e:
            print(f"加载 CSV 文件时出错：{e}")
            return {}

    def single_replace(self, text, target, replacement):
        """替换单个变量"""
        if not target or not replacement:
            return text
        target_clean = target.strip('"').strip()
        replacement_clean = replacement.strip('"').strip()
        return re.sub(re.escape(target_clean), replacement_clean, text)

    @classmethod
    def INPUT_TYPES(cls):
        # 加载 CSV 数据
        cls.video_data = cls.load_csv_data(cls.CSV_PATH)
        
        # 为各分类设置默认值
        if "编剧模板" not in cls.video_data:
            cls.video_data["编剧模板"] = {"默认模板": {'content': '', 'variables': ''}}
        if "视频主题" not in cls.video_data:
            cls.video_data["视频主题"] = {"默认主题": {'content': '普通场景', 'variables': ''}}
        if "镜头运镜" not in cls.video_data:
            cls.video_data["镜头运镜"] = {"默认运镜": {'content': '平稳镜头', 'variables': ''}}
        if "光影氛围" not in cls.video_data:
            cls.video_data["光影氛围"] = {"默认光影": {'content': '自然光', 'variables': ''}}
        if "视觉风格" not in cls.video_data:
            cls.video_data["视觉风格"] = {"默认风格": {'content': '写实风格', 'variables': ''}}
        if "音效" not in cls.video_data:
            cls.video_data["音效"] = {"默认音效": {'content': '背景音乐', 'variables': ''}}
        if "安全机制" not in cls.video_data:
            cls.video_data["安全机制"] = {"默认机制": {'content': '无特殊要求', 'variables': ''}}
        if "导演风格" not in cls.video_data:
            cls.video_data["导演风格"] = {"默认导演": {'content': '常规导演风格', 'variables': ''}}
        if "灵动变量" not in cls.video_data:
            cls.video_data["灵动变量"] = {"默认变量": {'content': '', 'variables': ''}}
        if "输出设置" not in cls.video_data:
            cls.video_data["输出设置"] = {"默认输出": {'content': '标准输出', 'variables': ''}}
        
        # 为每个类别添加"自定义"选项
        for category in cls.video_data:
            category_data = cls.video_data[category]
            category_data["自定义"] = {'content': '', 'variables': ''}
        
        return {
            "required": {
                "script_template": (list(cls.video_data.get("编剧模板", {}).keys()), {"label": "编剧模板"}),
                "custom_template": ("STRING", {"default": "", "multiline": True, "placeholder": "输入自定义模板（留空则使用选中模板）"}),
            },
            "optional": {
                "director": (list(cls.video_data.get("导演风格", {}).keys()), {"default": "自定义", }),
                "director_custom": ("STRING", {"default": "", "multiline": True, "placeholder": "输入自定义导演风格内容"}),
                "video_theme": (list(cls.video_data.get("视频主题", {}).keys()), {"default": "自定义", }),
                "video_theme_custom": ("STRING", {"default": "", "multiline": True, "placeholder": "输入自定义视频主题内容"}),
                "camera_movement": (list(cls.video_data.get("镜头运镜", {}).keys()), {"default": "自定义", }),
                "camera_custom": ("STRING", {"default": "", "multiline": True, "placeholder": "输入自定义镜头运镜内容"}),
                "lighting": (list(cls.video_data.get("光影氛围", {}).keys()), {"default": "自定义", }),
                "lighting_custom": ("STRING", {"default": "", "multiline": True, "placeholder": "输入自定义光影氛围内容"}),
                "visual_style": (list(cls.video_data.get("视觉风格", {}).keys()), {"default": "自定义", }),
                "style_custom": ("STRING", {"default": "", "multiline": True, "placeholder": "输入自定义视觉风格内容"}),
                "sound": (list(cls.video_data.get("音效", {}).keys()), {"default": "自定义", }),
                "sound_custom": ("STRING", {"default": "", "multiline": True, "placeholder": "输入自定义音效内容"}),
                "safety": (list(cls.video_data.get("安全机制", {}).keys()), {"default": "自定义", }),
                "safety_custom": ("STRING", {"default": "", "multiline": True, "placeholder": "输入自定义安全机制内容"}),
                "output_settings": (list(cls.video_data.get("输出设置", {}).keys()), {"default": "自定义", }),
                "output_custom": ("STRING", {"default": "", "multiline": True, "placeholder": "输入自定义输出设置内容"}),
                "dynamic_variable": (list(cls.video_data.get("灵动变量", {}).keys()), {"default": "自定义", }),
                "dynamic_custom": ("STRING", {"default": "", "multiline": True, "placeholder": "输入自定义灵动变量内容"}),

            }
        }
        
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("script",)
    FUNCTION = "generate_script"
    CATEGORY = "Apt_Preset/prompt"

    DESCRIPTION = """
    视频助手 - 通过变量替换生成编剧模板
    使用示例：
    1. 选择一个编剧模板（如：编剧 - 武术短片）
    2. 选择或自定义各项参数（视频主题、镜头运镜、光影氛围等）
    3. 节点会自动替换模板中的所有变量占位符
    4. 输出完整的编剧脚本
    自定义功能：
    - 可以在 custom_template 字段输入自定义模板
    - 每个参数都支持自定义值（如 video_theme_custom）
    - 下拉框选择的值优先于自定义值
    """

    def generate_script(self, script_template, 
                        video_theme="自定义", video_theme_custom="", 
                        camera_movement="自定义", camera_custom="", 
                        lighting="自定义", lighting_custom="", 
                        visual_style="自定义", style_custom="", 
                        sound="自定义", sound_custom="", 
                        safety="自定义", safety_custom="", 
                        director="自定义", director_custom="", 
                        output_settings="自定义", output_custom="", 
                        dynamic_variable="自定义", dynamic_custom="", 
                        custom_template=""):
        
        # 获取各项内容，优先使用下拉框选择的内容，如果选择的是"自定义"则使用自定义输入
        theme_content = self.video_data.get("视频主题", {}).get(video_theme, {}).get('content', '') if video_theme != "自定义" else video_theme_custom
        
        camera_content = self.video_data.get("镜头运镜", {}).get(camera_movement, {}).get('content', '') if camera_movement != "自定义" else camera_custom
        
        lighting_content = self.video_data.get("光影氛围", {}).get(lighting, {}).get('content', '') if lighting != "自定义" else lighting_custom
        
        style_content = self.video_data.get("视觉风格", {}).get(visual_style, {}).get('content', '') if visual_style != "自定义" else style_custom
        
        sound_content = self.video_data.get("音效", {}).get(sound, {}).get('content', '') if sound != "自定义" else sound_custom
        
        safety_content = self.video_data.get("安全机制", {}).get(safety, {}).get('content', '') if safety != "自定义" else safety_custom
        
        director_content = self.video_data.get("导演风格", {}).get(director, {}).get('content', '') if director != "自定义" else director_custom
        
        output_content = self.video_data.get("输出设置", {}).get(output_settings, {}).get('content', '') if output_settings != "自定义" else output_custom
        
        dynamic_content = self.video_data.get("灵动变量", {}).get(dynamic_variable, {}).get('content', '') if dynamic_variable != "自定义" else dynamic_custom
        
        # 获取模板内容
        if custom_template.strip():
            # 使用自定义模板
            template_content = custom_template.strip()
        else:
            # 使用选中的模板
            template_content = self.video_data.get("编剧模板", {}).get(script_template, {}).get('content', '')
            if not template_content:
                template_content = "❌ 未找到模板内容"
        
        # 如果模板内容为空，使用默认模板
        if not template_content or template_content == "❌ 未找到模板内容":
            template_content = "请选择有效的编剧模板或输入自定义模板。"
        
        # 执行变量替换
        replacements = {
            "{视频主题}": theme_content,
            "{镜头运镜}": camera_content,
            "{光影氛围}": lighting_content,
            "{视觉风格}": style_content,
            "{音效}": sound_content,
            "{导演风格}": director_content,
            "{输出设置}": output_content,
            "{安全机制}": safety_content,
            "{灵动变量}": dynamic_content
        }
        
        # 依次替换所有变量
        result = template_content
        for placeholder, value in replacements.items():
            if value:  # 只替换有内容的变量
                result = self.single_replace(result, placeholder, value)
        
        return (result,)
